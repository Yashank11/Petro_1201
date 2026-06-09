"""
World Bank Global Gas Flaring Data Loader
==========================================
Reads the two offline Excel files downloaded from:
  https://www.worldbank.org/en/programs/gasflaringreduction/global-flaring-data

Files (stored in backend/data/):
  wb_flaring_by_economy_2012_2024.xlsx  — annual BCM by country, 2012-2024
  wb_flaring_by_location_2012_2024.xlsx — per-flare-site BCM, 2012-2024

Public API:
  get_historical_trends()   → [{year, global_bcm, top_countries: [{country, bcm}]}, ...]
  get_country_history(name) → [{year, bcm}, ...]
  get_location_data(country, year) → [{lat, lon, bcm, field, operator, type}, ...]
"""

import os
import openpyxl

_DATA_DIR = os.path.dirname(__file__)
_ECONOMY_FILE  = os.path.join(_DATA_DIR, "wb_flaring_by_economy_2012_2024.xlsx")
_LOCATION_FILE = os.path.join(_DATA_DIR, "wb_flaring_by_location_2012_2024.xlsx")

# ── Country name normalisation (World Bank name → our app name) ───────────────
_NAME_MAP = {
    "Iran, Islamic Rep.":         "Iran",
    "Russian Federation":         "Russia",
    "United States":              "USA",
    "Venezuela, RB":              "Venezuela",
    "Saudi Arabia":               "Saudi Arabia",
    "Iraq":                       "Iraq",
    "Nigeria":                    "Nigeria",
    "Algeria":                    "Algeria",
    "Libya":                      "Libya",
    "Mexico":                     "Mexico",
    "Kazakhstan":                 "Kazakhstan",
    "Canada":                     "Canada",
    "India":                      "India",
    "Angola":                     "Angola",
    "Norway":                     "Norway",
    "Oman":                       "Oman",
    "United Arab Emirates":       "UAE",
    "Kuwait":                     "Kuwait",
    "Egypt, Arab Rep.":           "Egypt",
    "Malaysia":                   "Malaysia",
    "Indonesia":                  "Indonesia",
    "Brazil":                     "Brazil",
    "Argentina":                  "Argentina",
    "Ecuador":                    "Ecuador",
    "Colombia":                   "Colombia",
    "Gabon":                      "Gabon",
    "Congo, Dem. Rep.":           "DRC",
    "Congo, Rep.":                "Congo",
    "Sudan":                      "Sudan",
    "South Sudan":                "South Sudan",
    "Chad":                       "Chad",
    "Cameroon":                   "Cameroon",
    "Equatorial Guinea":          "Eq. Guinea",
    "Côte d'Ivoire":              "Côte d'Ivoire",
    "Ghana":                      "Ghana",
    "Turkmenistan":               "Turkmenistan",
    "Azerbaijan":                 "Azerbaijan",
    "Uzbekistan":                 "Uzbekistan",
    "Syrian Arab Republic":       "Syria",
    "Yemen, Rep.":                "Yemen",
    "Myanmar":                    "Myanmar",
    "Vietnam":                    "Vietnam",
    "Papua New Guinea":           "PNG",
    "Trinidad and Tobago":        "Trinidad",
    "Bolivia":                    "Bolivia",
    "Bahrain":                    "Bahrain",
    "Qatar":                      "Qatar",
}


def _normalise_country(raw: str) -> str:
    raw = str(raw).strip()
    return _NAME_MAP.get(raw, raw)


# ── Economy loader ────────────────────────────────────────────────────────────

def _load_economy() -> dict:
    """
    Returns {country_name: {year: bcm}} for all rows in the 'Flare volume' sheet.
    """
    if not os.path.exists(_ECONOMY_FILE):
        return {}

    wb = openpyxl.load_workbook(_ECONOMY_FILE, read_only=True, data_only=True)
    ws = wb["Flare volume"]

    # Row 1: ['Country, bcm', 2012, 2013, ..., 2024]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    years  = [h for h in header[1:] if isinstance(h, int)]

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        raw_name = str(row[0]).strip()
        # Skip aggregate rows
        if raw_name.lower() in ("total", "world", "global"):
            continue
        country = _normalise_country(raw_name)
        data[country] = {}
        for i, yr in enumerate(years):
            val = row[i + 1]
            data[country][yr] = round(float(val), 4) if val is not None else 0.0

    wb.close()
    return data


# ── Location loader ───────────────────────────────────────────────────────────

def _load_locations() -> list:
    """
    Returns a list of dicts from the location-level sheet.
    Columns: Country, Latitude, Longitude, bcm, MMscfd, Year,
             Field Type, Field Name, Field Operator, Location, Flare Level
    """
    if not os.path.exists(_LOCATION_FILE):
        return []

    wb  = openpyxl.load_workbook(_LOCATION_FILE, read_only=True, data_only=True)
    ws  = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)

    # Normalise header keys
    col = {str(h).strip().replace("  ", " "): i for i, h in enumerate(header) if h}

    records = []
    for row in rows_iter:
        try:
            country = _normalise_country(row[col["Country"]])
            lat     = row[col["Latitude"]]
            lon     = row[col["Longitude"]]
            bcm     = row[col["bcm"]]
            year    = row[col["Year"]]
            field   = row[col.get("Field Name", -1)] if "Field Name" in col else ""
            op      = row[col.get("Field Operator", -1)] if "Field Operator" in col else ""
            ftype   = row[col.get("Field Type", -1)] if "Field Type" in col else ""
            loc     = row[col.get("Location", -1)] if "Location" in col else ""
            level   = row[col.get("Flare Level", -1)] if "Flare Level" in col else ""

            if lat is None or lon is None or year is None:
                continue

            records.append({
                "country":  country,
                "lat":      round(float(lat), 6),
                "lon":      round(float(lon), 6),
                "bcm":      round(float(bcm or 0), 6),
                "year":     int(year),
                "field":    str(field or ""),
                "operator": str(op or ""),
                "type":     str(ftype or ""),
                "location": str(loc or ""),
                "level":    str(level or ""),
            })
        except Exception:
            continue

    wb.close()
    return records


# ── Module-level cache (loaded once at import) ────────────────────────────────

_ECONOMY:   dict = _load_economy()
_LOCATIONS: list = _load_locations()
YEARS = sorted({yr for country_data in _ECONOMY.values() for yr in country_data})


# ── Public API ────────────────────────────────────────────────────────────────

def get_historical_trends(top_n: int = 10) -> list:
    """
    Returns yearly global totals + top-N emitters per year.
    [{year, global_bcm, top_countries: [{country, bcm}]}, ...]
    """
    result = []
    for yr in YEARS:
        row_data = []
        for country, years_data in _ECONOMY.items():
            bcm = years_data.get(yr, 0.0)
            if bcm > 0:
                row_data.append({"country": country, "bcm": bcm})

        row_data.sort(key=lambda x: -x["bcm"])
        global_bcm = round(sum(r["bcm"] for r in row_data), 3)

        result.append({
            "year":          yr,
            "global_bcm":    global_bcm,
            "top_countries": row_data[:top_n],
        })

    return result


def get_country_history(country_name: str) -> list:
    """
    Returns [{year, bcm}] for a specific country (case-insensitive partial match).
    """
    # Try exact match first
    data = _ECONOMY.get(country_name)
    if data is None:
        # Case-insensitive fallback
        lower = country_name.lower()
        for k, v in _ECONOMY.items():
            if k.lower() == lower or lower in k.lower():
                data = v
                break

    if data is None:
        return []

    return [{"year": yr, "bcm": data.get(yr, 0.0)} for yr in YEARS]


def get_location_data(country: str = None, year: int = None) -> list:
    """
    Returns filtered location records.
    If country is None → all countries. If year is None → all years.
    """
    records = _LOCATIONS
    if country and country.lower() != "all":
        lower = country.lower()
        records = [r for r in records if r["country"].lower() == lower]
    if year:
        records = [r for r in records if r["year"] == year]
    return records


def get_all_countries() -> list:
    """Returns sorted list of countries available in the economy dataset."""
    return sorted(_ECONOMY.keys())


if __name__ == "__main__":
    print(f"Economy rows : {len(_ECONOMY)}")
    print(f"Location rows: {len(_LOCATIONS)}")
    print(f"Years        : {YEARS}")
    trends = get_historical_trends(5)
    for t in trends[-3:]:
        print(f"  {t['year']}: {t['global_bcm']} BCM — Top: {[c['country'] for c in t['top_countries']]}")
