"""
Petro — Real Well Database
Parses all 5 country Excel files + Saudi Aramco.xlsx into a unified KNOWN_WELLS list.
Handles both DMS string coordinates (e.g., 26°45' 31.82" N) and plain decimal floats.
Saudi Aramco entries from both files are merged and deduplicated by proximity (<5 km).
"""
import os
import re
import math
from typing import Optional

import openpyxl


# ── DMS Parser ───────────────────────────────────────────────────────────────

def _parse_dms(raw) -> Optional[float]:
    """Convert DMS string OR decimal to float degrees. Returns None on failure."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()

    # Normalise fancy quotes / degree symbols
    s = s.replace('\u00b0', '°').replace('\u2019', "'").replace('\u2018', "'")

    # Pattern: 26°45' 31.82" N  or  26° 26' 00" N  or  30° 11' 14.64" N
    m = re.match(
        r'(\d+)[°d]\s*(\d+)[\'`\u2019]\s*([\d.]+)[\""]?\s*([NSEWnsew]?)',
        s
    )
    if m:
        deg, mn, sec, direction = m.groups()
        val = float(deg) + float(mn) / 60 + float(sec) / 3600
        if direction.upper() in ('S', 'W'):
            val = -val
        return round(val, 6)

    # Fallback: try plain float
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# ── Haversine distance ────────────────────────────────────────────────────────

def _haversine_km(lat1, lon1, lat2, lon2) -> float:
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return R * 2 * math.asin(math.sqrt(a))


# ── Country lookup by source file ────────────────────────────────────────────

_SOURCE_COUNTRY = {
    'wells india.xlsx':      'India',
    'wells iran.xlsx':       'Iran',
    'wells iraq.xlsx':       'Iraq',
    'wells saudi arab.xlsx': 'Saudi Arabia',
    'wells usa.xlsx':        'USA',
    'Saudi Aramco.xlsx':     'Saudi Arabia',
}


# ── File list ─────────────────────────────────────────────────────────────────

_DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..') # project root

_EXCEL_FILES = [
    'wells india.xlsx',
    'wells iran.xlsx',
    'wells iraq.xlsx',
    'wells saudi arab.xlsx',
    'wells usa.xlsx',
    'Saudi Aramco.xlsx',
]


def _normalise_row(headers: list, row: tuple) -> dict:
    """Map arbitrary header names to canonical keys."""
    d = {str(h).strip(): v for h, v in zip(headers, row) if h}
    
    def _pick(*keys):
        for k in keys:
            if k in d and d[k]:
                return d[k]
        return ''

    company  = _pick('Company Name')
    name     = _pick(
        'Name of Well or Facility', 'Name of well', 'Name of well / platform / field',
        'Name of Well', 'Name of well / Facility', 'Well Name'
    )
    lat_raw  = _pick('Latitude')
    lon_raw  = _pick('Longitude')
    landmark = _pick('City/Major Landmark Nearby', 'City/major landmark nearby',
                     'City / Major Landmark Nearby')
    return dict(company=company, name=name, lat_raw=lat_raw, lon_raw=lon_raw,
                landmark=landmark)


def _load_file(filename: str) -> list:
    fpath = os.path.join(_DATA_DIR, filename)
    if not os.path.exists(fpath):
        return []
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    country = _SOURCE_COUNTRY.get(filename, 'Unknown')
    wells = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        r = _normalise_row(headers, row)
        lat = _parse_dms(r['lat_raw'])
        lon = _parse_dms(r['lon_raw'])
        if lat is None or lon is None:
            # Still include with null coords — shown as "unknown location"
            wells.append(dict(
                name=str(r['name']),
                company=str(r['company']),
                lat=None,
                lon=None,
                landmark=str(r['landmark']),
                country=country,
                source=filename,
                detected=False,
                has_coords=False,
            ))
            continue
        wells.append(dict(
            name=str(r['name']),
            company=str(r['company']),
            lat=round(lat, 6),
            lon=round(lon, 6),
            landmark=str(r['landmark']),
            country=country,
            source=filename,
            detected=False,
            has_coords=True,
        ))
    return wells


def _dedup_by_proximity(wells: list, threshold_km: float = 2.0) -> list:
    """Remove duplicate entries within threshold_km of each other (same company)."""
    kept = []
    for w in wells:
        if not w['has_coords']:
            kept.append(w)
            continue
        duplicate = False
        for k in kept:
            if not k['has_coords']:
                continue
            if k['company'] == w['company']:
                dist = _haversine_km(w['lat'], w['lon'], k['lat'], k['lon'])
                if dist < threshold_km:
                    duplicate = True
                    break
        if not duplicate:
            kept.append(w)
    return kept


# ── Build KNOWN_WELLS at import time ─────────────────────────────────────────

def _build_database() -> list:
    all_wells = []
    for fname in _EXCEL_FILES:
        all_wells.extend(_load_file(fname))

    # Dedup Saudi Aramco (from both files)
    sa_wells  = [w for w in all_wells if w['country'] == 'Saudi Arabia']
    other     = [w for w in all_wells if w['country'] != 'Saudi Arabia']
    sa_dedup  = _dedup_by_proximity(sa_wells, threshold_km=2.0)
    return other + sa_dedup


KNOWN_WELLS: list = _build_database()

# Build a quick-lookup list of only coord-valid wells for spatial matching
COORD_WELLS = [w for w in KNOWN_WELLS if w['has_coords']]


def find_nearest_well(lat: float, lon: float, radius_km: float = 5.0) -> Optional[dict]:
    """
    Returns the closest known well within radius_km of (lat, lon).
    Uses exact coordinates from the Excel files.
    """
    best = None
    best_dist = float('inf')
    for w in COORD_WELLS:
        d = _haversine_km(lat, lon, w['lat'], w['lon'])
        if d < radius_km and d < best_dist:
            best_dist = d
            best = w
    return best


if __name__ == '__main__':
    print(f'Total known wells: {len(KNOWN_WELLS)}')
    print(f'  With coordinates: {len(COORD_WELLS)}')
    by_country = {}
    for w in KNOWN_WELLS:
        by_country[w['country']] = by_country.get(w['country'], 0) + 1
    for c, n in sorted(by_country.items()):
        print(f'  {c}: {n}')
